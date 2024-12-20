import datetime

import openpyxl as pyxl


def change_date_hdg(
    ws: pyxl.worksheet, fmt: str = "m%Y-%m", top_row: int = 1
) -> pyxl.worksheet:
    for a_col in ws.iter_cols(
        min_col=1, max_col=ws.max_column, min_row=top_row, max_row=top_row
    ):
        for a_cell in a_col:
            val = a_cell.value
            if isinstance(val, datetime.date):
                val = val.strftime(fmt)
                a_cell.value = val
    return ws


def add_col(
    ws: pyxl.worksheet, top_row: int, col: int, heading: str, value
) -> pyxl.worksheet:
    ws.insert_cols(col)
    ws.cell(row=top_row, column=col).value = heading
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=top_row + 1, max_row=ws.max_row
    ):
        for a_cell in a_row:
            a_cell.value = value
    return ws


def add_compte_no(ws: pyxl.worksheet, heading: str = "compte_no") -> pyxl.worksheet:
    """Create column to convert compte_no_text to compte_no_no"""
    ws.insert_cols(1)
    ws.cell(1, 1).value = heading
    for a_row in ws.iter_rows(min_col=1, max_col=1, min_row=2, max_row=ws.max_row):
        for a_cell in a_row:
            val = ws.cell(a_cell.row, a_cell.column + 1).value
            val = str(val)
            try:
                a_cell.value = val.replace("-", "")
            except TypeError:
                msg = "'{val}' is an illegal value."
                raise TypeError(msg)
    return ws


def add_headings(ws: pyxl.worksheet, top_row: int, **headings) -> pyxl.worksheet:
    for heading, col in headings.items():
        ws.cell(row=top_row, column=col).value = heading
    return ws
