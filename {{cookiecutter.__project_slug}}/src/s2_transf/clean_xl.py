import datetime
import logging
import re

import openpyxl as pyxl
from rich.logging import RichHandler

logging.basicConfig(
    level=logging.DEBUG,
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler()],
)
log = logging.getLogger("rich")


def wsdims(ws: pyxl.worksheet, verbose: bool = True) -> dict[str, str | int]:
    # source: https://stackoverflow.com/questions/52076181/python-openpyxl-sheet-dimensions
    dims = ws.dimensions
    sc, sr, ec, er = pyxl.utils.cell.range_boundaries(dims)
    if verbose:
        msg = f"dims={dims}: start col={sc}, start row={sr}, end col={ec}, end row={er}"
        print(msg)
    out = {"dim": dims, "sc": sc, "sr": sr, "ec": ec, "er": er}
    return out


def match_cols(
    ws: pyxl.worksheet, row: int, left_col: int, pattern: re.Pattern
) -> list[int]:
    idx = []
    for a_col in ws.iter_cols(
        min_col=left_col, max_col=ws.max_column, min_row=row, max_row=row
    ):
        for a_cell in a_col:
            val = str(a_cell.value)
            # print(val)
            # print(pattern.match(val))
            if pattern.match(val) is not None:
                # print(f"Match at : {a_cell.col_idx}")
                idx.append(a_cell.col_idx)
    # print(idx)
    return idx


def match_blank_cols(
    ws: pyxl.worksheet, row: int, left_col: int, min_len: int = 1
) -> list[int]:
    assert min_len >= 1, f"min_len must be >= 1. It is {min_len}."
    idx = []
    for a_col in ws.iter_cols(
        min_col=left_col, max_col=ws.max_column, min_row=row, max_row=row
    ):
        for a_cell in a_col:
            if a_cell.value is None:
                idx.append(a_cell.col_idx)
            elif len(str(a_cell.value)) < min_len:
                idx.append(a_cell.col_idx)
            else:
                pass
    return idx


def del_cols(ws: pyxl.worksheet, idx: list[int]) -> pyxl.worksheet:
    """Delete columns from a worksheet.

    When deleting columns, starts from the rightmost column.
    Otherwise ws changes and the index is wrong. This function prevents that.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        idx (list[int]): List of row no to delete.

    Returns:
        pyxl.worksheet: Worksheet with deleted columns.
    """
    idx = sorted(idx, reverse=True)
    for i in idx:
        ws.delete_cols(i)
    return ws


def del_left_cols(ws: pyxl.worksheet, left_col: int) -> pyxl.worksheet:
    """Delete the left columns from a spreadsheet.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        left_col (int): Column no before which all columns will be deleted.

    Returns:
        pyxl.worksheet: Worksheet with deleted columns.
    """
    assert left_col >= 2, f"left_col must be >= 2. It is {left_col}."
    idx = [*range(1, left_col)]
    del_cols(ws, idx)  # make sure you use del_cols to reverse order
    return ws


def match_rows(
    ws: pyxl.worksheet, col: int, first_row: int, pattern: re.Pattern
) -> list[int]:
    idx = []
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=first_row, max_row=ws.max_row
    ):
        for a_cell in a_row:
            val = str(a_cell.value)
            # print(val)
            # print(pattern.match(val))
            if pattern.match(val) is not None:
                # print(f"Match at : {a_cell.row}")
                idx.append(a_cell.row)
    # print(idx)
    return idx


def match_blank_rows(
    ws: pyxl.worksheet, col: int, first_row: int, min_len: int = 1
) -> list[int]:
    assert min_len >= 1, f"min_len must be >= 1. It is {min_len}."
    idx = []
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=first_row, max_row=ws.max_row
    ):
        for a_cell in a_row:
            if a_cell.value is None:
                idx.append(a_cell.row)
            elif len(str(a_cell.value)) < min_len:
                idx.append(a_cell.row)
            else:
                pass
    # print(idx)
    return idx


def match_type_date_rows(ws: pyxl.worksheet, col: int, first_row: int) -> list[int]:
    idx = []
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=first_row, max_row=ws.max_row
    ):
        for a_cell in a_row:
            if isinstance(a_cell.value, datetime.datetime):
                idx.append(a_cell.row)
    # print(idx)
    return idx


def del_rows(ws: pyxl.worksheet, idx: list[int]) -> pyxl.worksheet:
    """Delete rows from a worksheet.

    When deleting rows, starts from the bottom, otherwise ws changes and
    the index is wrong. This function prevents that.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        idx (list[int]): List of row no to delete.

    Returns:
        pyxl.worksheet: Worksheet with deleted rows.
    """
    idx = sorted(idx, reverse=True)
    for i in idx:
        ws.delete_rows(i)
    return ws


def del_top_rows(ws: pyxl.worksheet, top_row: int) -> pyxl.worksheet:
    """Delete the top rows from a spreadsheet.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        top_row (int): Row no before which all rows will be deleted.

    Returns:
        pyxl.worksheet: Worksheet with deleted rows.
    """
    assert top_row >= 2, f"top_row must be >= 2. It is {top_row}."
    idx = [*range(1, top_row)]
    del_rows(ws, idx)  # make sure you use del_rows to reverse order
    return ws


def negate_match(idx: list[int], lower: int, upper: int) -> list[int]:
    all_idx = [i for i in range(lower, upper)]
    new_idx = idx.copy()
    new_idx = [i for i in all_idx if i not in new_idx]
    return new_idx
