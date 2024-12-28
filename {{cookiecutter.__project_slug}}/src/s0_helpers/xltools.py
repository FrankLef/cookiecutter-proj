import datetime
import logging
import re

import openpyxl as pyxl
from rich.logging import RichHandler

logging.basicConfig(
    level=logging.DEBUG,
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler()],
)
log = logging.getLogger("rich")


def wsdims(ws: pyxl.worksheet, verbose: bool = False) -> dict[str, str | int]:
    """Get the dimensions of the active range fom an excel worksheet.

    Args:
        ws (pyxl.worksheet): Excel worksheet.
        verbose (bool, optional): Print the details to console. Defaults to False.

    Source:
        https://stackoverflow.com/questions/52076181/python-openpyxl-sheet-dimensions

    Returns:
        dict[str, str | int]: Dictionary with details on the active range.
    """
    # source: https://stackoverflow.com/questions/52076181/python-openpyxl-sheet-dimensions
    dims = ws.dimensions
    lc, tr, rc, br = pyxl.utils.cell.range_boundaries(dims)
    if verbose:
        msg = (
            f"dims={dims}: left col={lc}, top row={tr}, right col={rc}, bottom row={br}"
        )
        print(msg)
    out = {"dim": dims, "lc": lc, "tr": tr, "rc": rc, "br": br}
    return out


def match_cols(
    ws: pyxl.worksheet, row: int, left_col: int, pattern: re.Pattern
) -> list[int]:
    """Find column number matching a regex pattern at a given row."""
    idx = []
    for a_col in ws.iter_cols(
        min_col=left_col, max_col=ws.max_column, min_row=row, max_row=row
    ):
        for a_cell in a_col:
            val = str(a_cell.value)
            if pattern.match(val) is not None:
                idx.append(a_cell.col_idx)
    return idx


def match_blank_cols(
    ws: pyxl.worksheet, row: int, left_col: int, min_len: int = 1
) -> list[int]:
    """Find columns with a blank cell at a given row."""
    assert min_len >= 1, f"min_len must be >= 1. It is {min_len}."
    idx = []
    for a_col in ws.iter_cols(
        min_col=left_col, max_col=ws.max_column, min_row=row, max_row=row
    ):
        for a_cell in a_col:
            if a_cell.value is None:
                idx.append(a_cell.col_idx)
            elif len(str(a_cell.value)) < min_len:
                idx.append(a_cell.col_idx)
            else:
                pass
    return idx


def del_cols(ws: pyxl.worksheet, idx: list[int]) -> pyxl.worksheet:
    """Delete columns from a worksheet.

    When deleting columns, starts from the rightmost column.
    Otherwise ws changes and the index is wrong. This function prevents that.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        idx (list[int]): List of row no to delete.

    Returns:
        pyxl.worksheet: Worksheet with deleted columns.
    """
    idx = sorted(idx, reverse=True)
    for i in idx:
        ws.delete_cols(i)
    return ws


def del_left_cols(ws: pyxl.worksheet, left_col: int) -> pyxl.worksheet:
    """Delete the left columns from a spreadsheet.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        left_col (int): Column no before which all columns will be deleted.

    Returns:
        pyxl.worksheet: Worksheet with deleted columns.
    """
    assert left_col >= 2, f"left_col must be >= 2. It is {left_col}."
    idx = [*range(1, left_col)]
    del_cols(ws, idx)  # make sure you use del_cols to reverse order
    return ws


def match_rows(
    ws: pyxl.worksheet, col: int, first_row: int, pattern: re.Pattern
) -> list[int]:
    """Find the row number matching a regex pattern at a given column."""
    idx = []
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=first_row, max_row=ws.max_row
    ):
        for a_cell in a_row:
            val = str(a_cell.value)
            # print(val)
            # print(pattern.match(val))
            if pattern.match(val) is not None:
                # print(f"Match at : {a_cell.row}")
                idx.append(a_cell.row)
    # print(idx)
    return idx


def match_blank_rows(
    ws: pyxl.worksheet, col: int, first_row: int, min_len: int = 1
) -> list[int]:
    """Find rows with a blank cell at a given column."""
    assert min_len >= 1, f"min_len must be >= 1. It is {min_len}."
    idx = []
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=first_row, max_row=ws.max_row
    ):
        for a_cell in a_row:
            if a_cell.value is None:
                idx.append(a_cell.row)
            elif len(str(a_cell.value)) < min_len:
                idx.append(a_cell.row)
            else:
                pass
    # print(idx)
    return idx


def match_type_date_rows(ws: pyxl.worksheet, col: int, first_row: int) -> list[int]:
    """Fin the row with a date at a given column."""
    idx = []
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=first_row, max_row=ws.max_row
    ):
        for a_cell in a_row:
            if isinstance(a_cell.value, datetime.datetime):
                idx.append(a_cell.row)
    # print(idx)
    return idx


def del_rows(ws: pyxl.worksheet, idx: list[int]) -> pyxl.worksheet:
    """Delete rows from a worksheet.

    When deleting rows, starts from the bottom, otherwise ws changes and
    the index is wrong. This function prevents that.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        idx (list[int]): List of row no to delete.

    Returns:
        pyxl.worksheet: Worksheet with deleted rows.
    """
    idx = sorted(idx, reverse=True)
    for i in idx:
        ws.delete_rows(i)
    return ws


def del_top_rows(ws: pyxl.worksheet, top_row: int) -> pyxl.worksheet:
    """Delete the top rows from a spreadsheet.

    Args:
        ws (pyxl.worksheet): Worksheet to process.
        top_row (int): Row no before which all rows will be deleted.

    Returns:
        pyxl.worksheet: Worksheet with deleted rows.
    """
    assert top_row >= 2, f"top_row must be >= 2. It is {top_row}."
    idx = [*range(1, top_row)]
    del_rows(ws, idx)  # make sure you use del_rows to reverse order
    return ws


def negate_match(idx: list[int], lower: int, upper: int) -> list[int]:
    """Reverse the index number.

    In particular, this means that we include the number in the list and include now the numbers that were originaly excluded.
    """
    all_idx = [i for i in range(lower, upper)]
    new_idx = idx.copy()
    new_idx = [i for i in all_idx if i not in new_idx]
    return new_idx


def change_date_hdg(
    ws: pyxl.worksheet, fmt: str = "m%Y-%m", top_row: int = 1
) -> pyxl.worksheet:
    """Change date heading to a given date format."""
    for a_col in ws.iter_cols(
        min_col=1, max_col=ws.max_column, min_row=top_row, max_row=top_row
    ):
        for a_cell in a_col:
            val = a_cell.value
            if isinstance(val, datetime.date):
                val = val.strftime(fmt)
                a_cell.value = val
    return ws


def add_col(
    ws: pyxl.worksheet, heading: str, value, top_row: int = 1, col: int = 1
) -> pyxl.worksheet:
    """Insert a column with values

    Args:
        ws (pyxl.worksheet): _description_
        heading (str): _description_
        value (any): _description_
        top_row (int, optional): _description_. Defaults to 1.
        col (int, optional): _description_. Defaults to 1.

    Returns:
        pyxl.worksheet: _description_
    """
    ws.insert_cols(col)
    ws.cell(row=top_row, column=col).value = heading
    for a_row in ws.iter_rows(
        min_col=col, max_col=col, min_row=top_row + 1, max_row=ws.max_row
    ):
        for a_cell in a_row:
            a_cell.value = value
    return ws


def add_headings(ws: pyxl.worksheet, top_row: int, **headings) -> pyxl.worksheet:
    for heading, col in headings.items():
        ws.cell(row=top_row, column=col).value = heading
    return ws
